import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import json

EXCEL_DIR = "/home/claude/project2/excel"
DATA = "/home/claude/project2/data"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # deep navy
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=10)     # blue = hardcoded input/assumption
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)   # black = formula
LINK_FONT = Font(name=FONT_NAME, color="008000", size=10)      # green = cross-sheet link
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="7F7F7F")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def style_header_row(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def autosize(ws, widths=None):
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    else:
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 45)


def write_df(ws, df, start_row=1, header=True, font=FORMULA_FONT, number_formats=None):
    number_formats = number_formats or {}
    if header:
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=start_row, column=j, value=col)
        style_header_row(ws, row=start_row, ncols=len(df.columns))
        start_row += 1
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row + i, column=j, value=row[col])
            cell.font = font
            cell.border = BORDER
            if col in number_formats:
                cell.number_format = number_formats[col]
    return start_row + len(df)  # next free row


# ============================================================================
# TAB: README
# ============================================================================
ws = wb.create_sheet("README")
ws.sheet_view.showGridLines = False
ws["B2"] = "Meridian Manufacturing Co. — Demand Forecasting & Inventory Balancing Model"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Supply Chain Analytics Portfolio — Project 2"
ws["B3"].font = SUBTITLE_FONT

readme_lines = [
    ("", ""),
    ("What this workbook does", ""),
    ("", "Turns a simulated 12-month NetSuite-style sales/inventory export (40 SKUs) into a live"),
    ("", "replenishment planning model: ABC/XYZ classification, safety stock & reorder points, EOQ,"),
    ("", "a FORECAST.LINEAR demand baseline with a measured MAPE, and a red/yellow/green inventory"),
    ("", "health flag for every SKU."),
    ("", ""),
    ("Tab guide", ""),
    ("Assumptions", "Every hardcoded input used anywhere in this model, in one place."),
    ("Data_Quality_Log", "What was messy in the raw export and how it was cleaned before this workbook."),
    ("Raw_SKU_Master", "40-SKU catalog: category, supplier, cost, price, UOM, lead time."),
    ("Raw_Sales_Weekly", "1,805 rows of cleaned weekly sell-through, one row per SKU per week."),
    ("Raw_Inventory_Monthly", "480 rows of month-end on-hand / on-order snapshots."),
    ("ABC_XYZ_Classification", "Revenue-based ABC tier x demand-variability XYZ tier for every SKU."),
    ("Inventory_Planning_Model", "Safety stock, reorder point, EOQ, and stockout/overstock status — all live formulas."),
    ("Forecast_Detail", "Week-by-week forecast-vs-actual for the 6-week holdout, 36 eligible SKUs."),
    ("Forecast_MAPE_Summary", "Forecast accuracy (MAPE) by SKU, plus the catalog-wide average."),
    ("Executive_Summary", "The one-page view: KPI cards, ABC/XYZ matrix, and the SKUs that matter most."),
    ("", ""),
    ("Color key", ""),
    ("Blue text", "Hardcoded input or assumption"),
    ("Black text", "Formula (same sheet)"),
    ("Green text", "Formula linking to another sheet"),
    ("Red / Yellow / Green fill", "Stockout risk / Overstock / Healthy (Inventory_Planning_Model tab)"),
]
r = 5
for label, text in readme_lines:
    if label and not text:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT_NAME, bold=True, size=11, color="1F3864")
    elif label and text:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
        ws.cell(row=r, column=3, value=text).font = Font(name=FONT_NAME, size=10)
    elif text:
        ws.cell(row=r, column=3, value=text).font = Font(name=FONT_NAME, size=10)
    r += 1
autosize(ws, widths={"A": 3, "B": 26, "C": 78})

# ============================================================================
# TAB: Assumptions
# ============================================================================
ws = wb.create_sheet("Assumptions")
ws.sheet_view.showGridLines = False
ws["B2"] = "Model Assumptions"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Every hardcoded number used in a formula elsewhere in this workbook is defined here, once."
ws["B3"].font = SUBTITLE_FONT

assumptions = [
    ("Name", "Value", "Cell", "Used for", "Rationale"),
    ("Service_Level_Z", 1.65, "C6", "Safety stock", "Z-score for a 95% single-tail service level (industry-standard planning target)."),
    ("Order_Cost_S", 75, "C7", "EOQ", "Assumed fixed cost per purchase order ($): PO processing + receiving labor."),
    ("Holding_Cost_Rate", 0.22, "C8", "EOQ", "Annual carrying cost as % of unit cost (typical industrial range: 18-30%)."),
    ("MOQ_Lot_Size", 25, "C9", "Order rounding", "Assumed minimum order lot size (units) enforced via ROUNDUP across the catalog."),
    ("Forecast_Holdout_Weeks", 6, "C10", "MAPE test", "Most recent 6 weeks per SKU held out and forecast from the prior weeks to measure accuracy."),
    ("Min_Weeks_For_Forecast", 15, "C11", "MAPE eligibility", "SKUs with fewer weeks of history are excluded from statistical forecasting (too new/sparse)."),
    ("ABC_A_Cutoff", 0.80, "C12", "ABC class", "SKUs contributing to the top 80% of cumulative revenue -> Class A."),
    ("ABC_B_Cutoff", 0.95, "C13", "ABC class", "Next slice, 80-95% cumulative revenue -> Class B. Remainder -> Class C."),
    ("XYZ_X_Cutoff", 0.30, "C14", "XYZ class", "Coefficient of variation <= 0.30 -> steady demand (X)."),
    ("XYZ_Y_Cutoff", 0.55, "C15", "XYZ class", "CV 0.30-0.55 -> moderate variability (Y). Above 0.55 -> erratic (Z)."),
    ("Stockout_Multiple", 1.0, "C16", "Inventory status", "Days of supply (incl. pipeline) < 1x lead time -> Stockout Risk."),
    ("Overstock_Multiple", 3.0, "C17", "Inventory status", "Days of supply (incl. pipeline) > 3x lead time -> Overstock."),
    ("Weeks_Per_Year", 52, "C18", "Annualization", "Used to annualize weekly demand into EOQ's D term."),
]
r = 5
for row in assumptions:
    for j, val in enumerate(row, start=2):
        cell = ws.cell(row=r, column=j, value=val)
        cell.border = BORDER
        if r == 5:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.font = INPUT_FONT if j == 3 else Font(name=FONT_NAME, size=10)
            if j == 3 and isinstance(val, float) and val < 1:
                cell.number_format = "0%"
    r += 1
autosize(ws, widths={"A": 3, "B": 22, "C": 10, "D": 8, "E": 20, "F": 62})
ws.row_dimensions[5].height = 26

# ============================================================================
# TAB: Data_Quality_Log
# ============================================================================
ws = wb.create_sheet("Data_Quality_Log")
ws.sheet_view.showGridLines = False
ws["B2"] = "Data Quality Log — Raw NetSuite Export"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Issues found in the raw export and how each was resolved before building this workbook."
ws["B3"].font = SUBTITLE_FONT

dq_rows = [
    ("Issue", "Where found", "Raw example", "Resolution"),
    ("Mixed date formats", "Sales export, Order_Date", "01/18/2026, 2026-01-18, 18-Jan-2026, 01/18/26",
     "Parsed all 4 formats to ISO dates in Python/SQL before loading; 0 unparseable after cleaning."),
    ("Duplicate transaction rows", "Sales export", "27 exact-duplicate rows (~1.5% of file)",
     "Dropped via SQL DISTINCT on SKU + date + qty + revenue + warehouse."),
    ("UOM inconsistency", "SKU master & inventory snapshot", "EA / Each / each / ea.  |  PCS / pcs / Pcs",
     "Standardized to 4 canonical codes (EA, PCS, BOX, FT) via a lookup table."),
    ("Supplier name casing", "SKU master", "Midwest Fastener Co. / MIDWEST FASTENER CO. / midwest fastener co.",
     "Standardized to title case."),
    ("Missing On_Order_Qty", "Inventory snapshot", "46 of 480 rows (9.6%) blank",
     "Treated as 0 (no open PO) with a flag column retained so the gap stays visible, not hidden."),
    ("Sparse / new-SKU history", "Sales export", "4 SKUs with <15 weeks of recorded sales",
     "Excluded from statistical forecasting (Forecast_MAPE_Summary); flagged for judgmental forecasting instead."),
]
r = 5
for row in dq_rows:
    for j, val in enumerate(row, start=2):
        cell = ws.cell(row=r, column=j, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r == 5:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.font = Font(name=FONT_NAME, size=10)
    ws.row_dimensions[r].height = 30 if r > 5 else 26
    r += 1
autosize(ws, widths={"A": 3, "B": 24, "C": 20, "D": 34, "E": 55})

# ============================================================================
# TAB: Raw_SKU_Master
# ============================================================================
sku = pd.read_csv(f"{DATA}/sku_master_clean.csv")
sku_out = sku[["SKU_ID", "SKU_Description", "Category", "Supplier_Clean", "Unit_Cost",
               "Unit_Price", "UOM", "UOM_Clean", "Lead_Time_Days", "Warehouse"]].rename(
    columns={"Supplier_Clean": "Supplier (cleaned)", "UOM": "UOM (raw)", "UOM_Clean": "UOM (clean)"})
ws = wb.create_sheet("Raw_SKU_Master")
next_row = write_df(ws, sku_out, number_formats={"Unit_Cost": "$#,##0.00", "Unit_Price": "$#,##0.00"})
autosize(ws)
ws.freeze_panes = "A2"

# ============================================================================
# TAB: Raw_Sales_Weekly  (sorted by SKU, week -- contiguous ranges per SKU)
# ============================================================================
weekly = pd.read_csv(f"{EXCEL_DIR}/weekly_sales_sorted.csv")
ws = wb.create_sheet("Raw_Sales_Weekly")
next_row = write_df(ws, weekly, number_formats={"Revenue": "$#,##0.00"})
autosize(ws)
ws.freeze_panes = "A2"

# ============================================================================
# TAB: Raw_Inventory_Monthly (sorted) -- with an in-sheet formula-driven
# "On_Order_Qty_Filled" column that treats blank as 0 (visible cleaning logic)
# ============================================================================
inv = pd.read_csv(f"{EXCEL_DIR}/inventory_sorted.csv")
inv_out = inv[["SKU_ID", "Month", "On_Hand_Qty", "On_Order_Qty", "UOM", "Warehouse"]].copy()
ws = wb.create_sheet("Raw_Inventory_Monthly")
next_row = write_df(ws, inv_out)
# add the formula column
n_rows = len(inv_out)
ws.cell(row=1, column=7, value="On_Order_Qty_Filled")
style_header_row(ws, row=1, ncols=7)
for i in range(n_rows):
    r = 2 + i
    cell = ws.cell(row=r, column=7, value=f"=IF(D{r}=\"\",0,D{r})")
    cell.font = FORMULA_FONT
    cell.border = BORDER
autosize(ws)
ws.freeze_panes = "A2"

wb.save(f"{EXCEL_DIR}/Project2_Demand_Forecasting_Inventory_Model.xlsx")
print(f"Stage 1 saved. Raw_SKU_Master: {len(sku_out)} rows | "
      f"Raw_Sales_Weekly: {len(weekly)} rows | Raw_Inventory_Monthly: {n_rows} rows")
