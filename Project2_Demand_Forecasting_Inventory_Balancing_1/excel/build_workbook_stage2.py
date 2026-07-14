import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

EXCEL_DIR = "/home/claude/project2/excel"
FPATH = f"{EXCEL_DIR}/Project2_Demand_Forecasting_Inventory_Model.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)
LINK_FONT = Font(name=FONT_NAME, color="008000", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.load_workbook(FPATH)

sku_by_rev = pd.read_csv(f"{EXCEL_DIR}/sku_by_revenue.csv")
sku_order = sku_by_rev["SKU_ID"].tolist()  # presentation order only; every value below is a live formula
assert len(sku_order) == 40

RSW_LAST = 1806   # Raw_Sales_Weekly data rows 2..1806
RSM_LAST = 41     # Raw_SKU_Master data rows 2..41

ws = wb.create_sheet("ABC_XYZ_Classification")
ws.sheet_view.showGridLines = False
ws["B2"] = "ABC (Revenue) x XYZ (Demand Variability) Classification"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Sorted by revenue, descending. Every value is a live formula against Raw_Sales_Weekly / Raw_SKU_Master."
ws["B3"].font = SUBTITLE_FONT

headers = ["SKU_ID", "SKU_Description", "Category", "Total_Revenue", "Total_Qty_Sold",
           "Cum_Revenue", "Cum_Pct", "ABC_Class", "Mean_Weekly_Qty", "StDev_Weekly_Qty",
           "CV", "XYZ_Class", "ABC_XYZ"]
HEADER_ROW = 5
DATA_START = 6
for j, h in enumerate(headers, start=1):
    ws.cell(row=HEADER_ROW, column=j, value=h)
style = lambda c: (c.fill == None)  # placeholder (unused)
for j in range(1, len(headers) + 1):
    c = ws.cell(row=HEADER_ROW, column=j)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[HEADER_ROW].height = 26

for i, sid in enumerate(sku_order):
    r = DATA_START + i
    ws.cell(row=r, column=1, value=sid).font = FORMULA_FONT  # SKU_ID (from pre-sorted list)

    # Description / Category via INDEX/MATCH against Raw_SKU_Master
    ws.cell(row=r, column=2,
            value=f"=INDEX(Raw_SKU_Master!$B$2:$B${RSM_LAST},MATCH(A{r},Raw_SKU_Master!$A$2:$A${RSM_LAST},0))")
    ws.cell(row=r, column=3,
            value=f"=INDEX(Raw_SKU_Master!$C$2:$C${RSM_LAST},MATCH(A{r},Raw_SKU_Master!$A$2:$A${RSM_LAST},0))")

    # Revenue & Qty via SUMIF against Raw_Sales_Weekly
    ws.cell(row=r, column=4,
            value=f"=SUMIF(Raw_Sales_Weekly!$A$2:$A${RSW_LAST},A{r},Raw_Sales_Weekly!$D$2:$D${RSW_LAST})")
    ws.cell(row=r, column=5,
            value=f"=SUMIF(Raw_Sales_Weekly!$A$2:$A${RSW_LAST},A{r},Raw_Sales_Weekly!$C$2:$C${RSW_LAST})")

    # Cumulative revenue & % (rows are pre-sorted descending by revenue)
    ws.cell(row=r, column=6, value=f"=SUM($D${DATA_START}:D{r})")
    ws.cell(row=r, column=7, value=f"=F{r}/SUM($D${DATA_START}:$D${DATA_START+39})")

    # ABC class from Assumptions cutoffs
    ws.cell(row=r, column=8,
            value=f"=IF(G{r}<=Assumptions!$C$12,\"A\",IF(G{r}<=Assumptions!$C$13,\"B\",\"C\"))")

    # Mean / StDev / CV of weekly qty (population stdev via SUMPRODUCT, no CSE array needed)
    ws.cell(row=r, column=9,
            value=f"=AVERAGEIF(Raw_Sales_Weekly!$A$2:$A${RSW_LAST},A{r},Raw_Sales_Weekly!$C$2:$C${RSW_LAST})")
    ws.cell(row=r, column=10,
            value=(f"=SQRT(SUMPRODUCT((Raw_Sales_Weekly!$A$2:$A${RSW_LAST}=A{r})*"
                   f"(Raw_Sales_Weekly!$C$2:$C${RSW_LAST}-I{r})^2)/"
                   f"COUNTIF(Raw_Sales_Weekly!$A$2:$A${RSW_LAST},A{r}))"))
    ws.cell(row=r, column=11, value=f"=IFERROR(J{r}/I{r},0)")

    # XYZ class from Assumptions cutoffs
    ws.cell(row=r, column=12,
            value=f"=IF(K{r}<=Assumptions!$C$14,\"X\",IF(K{r}<=Assumptions!$C$15,\"Y\",\"Z\"))")
    ws.cell(row=r, column=13, value=f"=H{r}&L{r}")

    for j in range(1, 14):
        c = ws.cell(row=r, column=j)
        c.font = FORMULA_FONT if j not in (2, 3) else LINK_FONT
        c.border = BORDER
        if j == 4 or j == 6:
            c.number_format = "$#,##0"
        elif j == 7:
            c.number_format = "0.0%"
        elif j in (9, 10):
            c.number_format = "0.0"
        elif j == 11:
            c.number_format = "0.00"

# column widths
widths = {"A": 10, "B": 34, "C": 22, "D": 13, "E": 13, "F": 13, "G": 10,
          "H": 9, "I": 12, "J": 12, "K": 8, "L": 9, "M": 10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A6"

wb.save(FPATH)
print("Stage 2 (ABC_XYZ_Classification) saved.")
