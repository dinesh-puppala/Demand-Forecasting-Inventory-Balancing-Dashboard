import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
import pandas as pd
import json

EXCEL_DIR = "/home/claude/project2/excel"
FPATH = f"{EXCEL_DIR}/Project2_Demand_Forecasting_Inventory_Model.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)
LINK_FONT = Font(name=FONT_NAME, color="008000", size=10)
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RED_FILL = PatternFill("solid", fgColor="F8CBAD")
YELLOW_FILL = PatternFill("solid", fgColor="FFE699")
GREEN_FILL = PatternFill("solid", fgColor="C6E0B4")

wb = openpyxl.load_workbook(FPATH)

sku_by_rev = pd.read_csv(f"{EXCEL_DIR}/sku_by_revenue.csv")
sku_order = sku_by_rev["SKU_ID"].tolist()
with open(f"{EXCEL_DIR}/sku_row_ranges.json") as f:
    row_ranges = json.load(f)

RSM_LAST = 41
ABC_LAST = 45  # ABC_XYZ_Classification data rows 6..45

if "Inventory_Planning_Model" in wb.sheetnames:
    del wb["Inventory_Planning_Model"]
ws = wb.create_sheet("Inventory_Planning_Model")
ws.sheet_view.showGridLines = False
ws["B2"] = "Inventory Planning Model"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Safety stock, reorder point, EOQ, and next-period forecast -- every cell is a live formula."
ws["B3"].font = SUBTITLE_FONT

HEADER_ROW = 8
DATA_START = 9
DATA_END = DATA_START + 39  # 48

ws.cell(row=5, column=2, value="Excess-Risk Days-of-Supply Threshold (75th pctile, floor 60 days):").font = LABEL_FONT
ws.cell(row=5, column=6, value=f"=MAX(60,PERCENTILE(P{DATA_START}:P{DATA_END},0.75))").font = Font(name=FONT_NAME, bold=True, color="C00000", size=10)
ws.cell(row=5, column=6).number_format = "0.0"

headers = ["SKU_ID", "SKU_Description", "Category", "Lead_Time_Days", "Unit_Cost",
           "Avg_Daily_Usage", "StDev_Weekly_Qty", "Lead_Time_Weeks", "Safety_Stock",
           "Reorder_Point", "Annual_Demand_Qty", "EOQ", "EOQ_Rounded_to_MOQ",
           "Avg_On_Hand_Qty", "Avg_On_Order_Qty", "Days_of_Supply",
           "Days_of_Supply_Incl_Pipeline", "Inventory_Status", "Next_Period_Forecast_Qty",
           "XYZ_Class", "ABC_Class", "Excess_Risk_SKU", "Avg_Inventory_Value"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=HEADER_ROW, column=j, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[HEADER_ROW].height = 30

for i, sid in enumerate(sku_order):
    r = DATA_START + i
    abc_row = 6 + i  # ABC_XYZ_Classification uses the SAME sku_order, same row offset (data starts row 6 there)

    ws.cell(row=r, column=1, value=sid).font = FORMULA_FONT
    ws.cell(row=r, column=2, value=f"=ABC_XYZ_Classification!B{abc_row}").font = LINK_FONT
    ws.cell(row=r, column=3, value=f"=ABC_XYZ_Classification!C{abc_row}").font = LINK_FONT
    ws.cell(row=r, column=4,
            value=f"=INDEX(Raw_SKU_Master!$I$2:$I${RSM_LAST},MATCH(A{r},Raw_SKU_Master!$A$2:$A${RSM_LAST},0))")
    ws.cell(row=r, column=5,
            value=f"=INDEX(Raw_SKU_Master!$E$2:$E${RSM_LAST},MATCH(A{r},Raw_SKU_Master!$A$2:$A${RSM_LAST},0))")

    # Avg daily usage from ABC_XYZ tab's Total_Qty_Sold (col E there), annualized via Assumptions weeks/year
    ws.cell(row=r, column=6, value=f"=ABC_XYZ_Classification!E{abc_row}/(Assumptions!$C$18*7)")
    ws.cell(row=r, column=7, value=f"=ABC_XYZ_Classification!J{abc_row}")
    ws.cell(row=r, column=8, value=f"=D{r}/7")

    # Safety stock, reorder point
    ws.cell(row=r, column=9, value=f"=ROUNDUP(Assumptions!$C$6*G{r}*SQRT(H{r}),0)")
    ws.cell(row=r, column=10, value=f"=ROUNDUP(F{r}*D{r}+I{r},0)")

    # EOQ
    ws.cell(row=r, column=11, value=f"=ABC_XYZ_Classification!E{abc_row}")
    ws.cell(row=r, column=12, value=f"=ROUND(SQRT((2*K{r}*Assumptions!$C$7)/(E{r}*Assumptions!$C$8)),0)")
    ws.cell(row=r, column=13, value=f"=ROUNDUP(L{r}/Assumptions!$C$9,0)*Assumptions!$C$9")

    # Inventory position from Raw_Inventory_Monthly (480 rows, cols A=SKU_ID, C=On_Hand, G=On_Order_Filled)
    ws.cell(row=r, column=14, value=f"=AVERAGEIF(Raw_Inventory_Monthly!$A$2:$A$481,A{r},Raw_Inventory_Monthly!$C$2:$C$481)")
    ws.cell(row=r, column=15, value=f"=AVERAGEIF(Raw_Inventory_Monthly!$A$2:$A$481,A{r},Raw_Inventory_Monthly!$G$2:$G$481)")

    ws.cell(row=r, column=16, value=f"=IFERROR(N{r}/F{r},0)")
    ws.cell(row=r, column=17, value=f"=IFERROR((N{r}+O{r})/F{r},0)")
    ws.cell(row=r, column=18,
            value=(f"=IF(Q{r}<Assumptions!$C$16*D{r},\"Stockout Risk\","
                   f"IF(Q{r}>Assumptions!$C$17*D{r},\"Overstock\",\"Healthy\"))"))

    # Next-period forecast: FORECAST.LINEAR over this SKU's own contiguous block in Raw_Sales_Weekly
    rr = row_ranges[sid]
    x_range = f"Raw_Sales_Weekly!$E${rr['start_row']}:$E${rr['end_row']}"
    y_range = f"Raw_Sales_Weekly!$C${rr['start_row']}:$C${rr['end_row']}"
    ws.cell(row=r, column=19,
            value=f"=ROUND(FORECAST(MAX({x_range})+1,{y_range},{x_range}),0)")

    ws.cell(row=r, column=20, value=f"=ABC_XYZ_Classification!L{abc_row}")
    ws.cell(row=r, column=21, value=f"=ABC_XYZ_Classification!H{abc_row}")
    ws.cell(row=r, column=22, value=f"=IF(AND(P{r}>$F$5,T{r}=\"Z\"),\"Yes\",\"No\")")
    ws.cell(row=r, column=23, value=f"=N{r}*E{r}")

    for j in range(1, 24):
        c = ws.cell(row=r, column=j)
        c.border = BORDER
        if c.font is None or c.font.name != FONT_NAME:
            pass
        if j not in (2, 3):
            c.font = FORMULA_FONT if j not in (2, 3) else LINK_FONT
        if j in (5,):
            c.number_format = "$#,##0.00"
        if j in (6, 7, 9, 10, 11, 12, 13, 14, 15):
            c.number_format = "#,##0.0"
        if j in (16, 17):
            c.number_format = "#,##0.0"
        if j == 19:
            c.number_format = "#,##0"
        if j == 23:
            c.number_format = "$#,##0"

# conditional formatting on Inventory_Status (col R = 18)
status_range = f"R{DATA_START}:R{DATA_END}"
ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Stockout Risk"'], fill=RED_FILL))
ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Overstock"'], fill=YELLOW_FILL))
ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Healthy"'], fill=GREEN_FILL))

widths = {"A": 10, "B": 30, "C": 20, "D": 8, "E": 9, "F": 11, "G": 10, "H": 9,
          "I": 10, "J": 11, "K": 11, "L": 8, "M": 11, "N": 11, "O": 11, "P": 10,
          "Q": 11, "R": 13, "S": 11, "T": 8, "U": 8, "V": 11, "W": 12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = f"A{DATA_START}"

wb.save(FPATH)
print("Stage 3 (Inventory_Planning_Model) saved.")
