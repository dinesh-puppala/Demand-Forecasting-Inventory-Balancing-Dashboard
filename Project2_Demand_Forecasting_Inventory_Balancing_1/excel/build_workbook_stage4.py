import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import json

EXCEL_DIR = "/home/claude/project2/excel"
FPATH = f"{EXCEL_DIR}/Project2_Demand_Forecasting_Inventory_Model.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)
LINK_FONT = Font(name=FONT_NAME, color="008000", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.load_workbook(FPATH)
for name in ("Forecast_Detail", "Forecast_MAPE_Summary"):
    if name in wb.sheetnames:
        del wb[name]

sku_by_rev = pd.read_csv(f"{EXCEL_DIR}/sku_by_revenue.csv")
sku_order = sku_by_rev["SKU_ID"].tolist()
with open(f"{EXCEL_DIR}/sku_row_ranges.json") as f:
    row_ranges = json.load(f)

eligible_skus = [s for s in sku_order if row_ranges[s]["eligible"]]
ineligible_skus = [s for s in sku_order if not row_ranges[s]["eligible"]]

# ============================================================================
# TAB: Forecast_Detail  (one row per SKU per holdout week)
# ============================================================================
ws = wb.create_sheet("Forecast_Detail")
ws.sheet_view.showGridLines = False
ws["B2"] = "Forecast Holdout Detail"
ws["B2"].font = TITLE_FONT
ws["B3"] = ("Most recent 6 weeks per SKU held out; each is forecast from FORECAST() over that SKU's prior "
            "weeks only (no leakage), then compared to what actually happened.")
ws["B3"].font = SUBTITLE_FONT

headers = ["SKU_ID", "Week_Index", "Actual_Qty", "Forecast_Qty", "APE", "Abs_Error"]
HEADER_ROW = 5
DATA_START = 6
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=HEADER_ROW, column=j, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

r = DATA_START
for sid in eligible_skus:
    rr = row_ranges[sid]
    train_x = f"Raw_Sales_Weekly!$E${rr['train_start_row']}:$E${rr['train_end_row']}"
    train_y = f"Raw_Sales_Weekly!$C${rr['train_start_row']}:$C${rr['train_end_row']}"
    for test_row in range(rr["test_start_row"], rr["test_end_row"] + 1):
        ws.cell(row=r, column=1, value=sid).font = FORMULA_FONT
        ws.cell(row=r, column=2, value=f"=Raw_Sales_Weekly!E{test_row}").font = LINK_FONT
        ws.cell(row=r, column=3, value=f"=Raw_Sales_Weekly!C{test_row}").font = LINK_FONT
        ws.cell(row=r, column=4,
                value=f"=ROUND(FORECAST(B{r},{train_y},{train_x}),0)").font = FORMULA_FONT
        ws.cell(row=r, column=5, value=f"=IFERROR(ABS((D{r}-C{r})/C{r}),\"\")").font = FORMULA_FONT
        ws.cell(row=r, column=6, value=f"=ABS(D{r}-C{r})").font = FORMULA_FONT
        for j in range(1, 7):
            ws.cell(row=r, column=j).border = BORDER
            if j == 5:
                ws.cell(row=r, column=j).number_format = "0.0%"
        r += 1
DETAIL_LAST_ROW = r - 1

widths = {"A": 10, "B": 11, "C": 11, "D": 12, "E": 9, "F": 11}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = f"A{DATA_START}"

# ============================================================================
# TAB: Forecast_MAPE_Summary
# ============================================================================
ws = wb.create_sheet("Forecast_MAPE_Summary")
ws.sheet_view.showGridLines = False
ws["B2"] = "Forecast Accuracy Summary (MAPE)"
ws["B2"].font = TITLE_FONT
ws["B3"] = "MAPE = mean absolute percentage error across each SKU's 6-week holdout. Lower is better."
ws["B3"].font = SUBTITLE_FONT

ws.cell(row=5, column=2, value="Catalog-wide MAPE (all holdout weeks, all eligible SKUs):").font = Font(name=FONT_NAME, bold=True, size=10)
ws.cell(row=5, column=6, value=f"=AVERAGE(Forecast_Detail!E{DATA_START}:E{DETAIL_LAST_ROW})").font = Font(name=FONT_NAME, bold=True, color="C00000", size=11)
ws.cell(row=5, column=6).number_format = "0.0%"

ws.cell(row=6, column=2, value="Catalog-wide WAPE (volume-weighted, robust to near-zero weeks):").font = Font(name=FONT_NAME, bold=True, size=10)
ws.cell(row=6, column=6, value=f"=SUM(Forecast_Detail!F{DATA_START}:F{DETAIL_LAST_ROW})/SUM(Forecast_Detail!C{DATA_START}:C{DETAIL_LAST_ROW})").font = Font(name=FONT_NAME, bold=True, color="1F7A1F", size=11)
ws.cell(row=6, column=6).number_format = "0.0%"

ws.cell(row=7, column=2,
        value=("Note: naive MAPE is inflated by a handful of holdout weeks with very low actual demand "
               "(a single-digit actual turns a small unit error into a triple-digit % error). WAPE weights "
               "every week by its volume and is the more representative accuracy figure for this catalog.")
        ).font = Font(name=FONT_NAME, italic=True, size=9, color="7F7F7F")

headers = ["SKU_ID", "SKU_Description", "Weeks_of_History", "Eligible_for_Forecast", "MAPE", "WAPE", "Accuracy_Tier"]
HEADER_ROW = 8
DATA_START2 = 9
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=HEADER_ROW, column=j, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[HEADER_ROW].height = 26

for i, sid in enumerate(sku_order):
    r = DATA_START2 + i
    abc_row = 6 + i
    rr = row_ranges[sid]
    ws.cell(row=r, column=1, value=sid).font = FORMULA_FONT
    ws.cell(row=r, column=2, value=f"=ABC_XYZ_Classification!B{abc_row}").font = LINK_FONT
    ws.cell(row=r, column=3, value=rr["n_weeks"]).font = FORMULA_FONT
    ws.cell(row=r, column=4, value="Yes" if rr["eligible"] else "No").font = FORMULA_FONT
    if rr["eligible"]:
        ws.cell(row=r, column=5,
                value=f"=AVERAGEIF(Forecast_Detail!$A${DATA_START}:$A${DETAIL_LAST_ROW},A{r},Forecast_Detail!$E${DATA_START}:$E${DETAIL_LAST_ROW})")
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=6,
                value=(f"=SUMIF(Forecast_Detail!$A${DATA_START}:$A${DETAIL_LAST_ROW},A{r},Forecast_Detail!$F${DATA_START}:$F${DETAIL_LAST_ROW})"
                       f"/SUMIF(Forecast_Detail!$A${DATA_START}:$A${DETAIL_LAST_ROW},A{r},Forecast_Detail!$C${DATA_START}:$C${DETAIL_LAST_ROW})"))
        ws.cell(row=r, column=6).number_format = "0.0%"
        ws.cell(row=r, column=7,
                value=f"=IF(F{r}<=0.15,\"Good\",IF(F{r}<=0.3,\"Fair\",\"Poor\"))")
    else:
        ws.cell(row=r, column=5, value="N/A")
        ws.cell(row=r, column=6, value="N/A")
        ws.cell(row=r, column=7, value="Insufficient history")
    for j in range(1, 8):
        ws.cell(row=r, column=j).border = BORDER

widths = {"A": 10, "B": 32, "C": 14, "D": 15, "E": 9, "F": 9, "G": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = f"A{DATA_START2}"

wb.save(FPATH)
print(f"Stage 4 saved. Forecast_Detail rows: {DATA_START}-{DETAIL_LAST_ROW} "
      f"({DETAIL_LAST_ROW-DATA_START+1} rows, {len(eligible_skus)} eligible SKUs x 6 weeks)")
print(f"Ineligible SKUs: {ineligible_skus}")
