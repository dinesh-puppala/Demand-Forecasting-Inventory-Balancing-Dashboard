import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

EXCEL_DIR = "/home/claude/project2/excel"
FPATH = f"{EXCEL_DIR}/Project2_Demand_Forecasting_Inventory_Model.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=10, color="595959")
KPI_VALUE_FONT = Font(name=FONT_NAME, bold=True, size=22, color="1F3864")
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CARD_FILL = PatternFill("solid", fgColor="F2F2F2")
RED_FILL = PatternFill("solid", fgColor="F8CBAD")
YELLOW_FILL = PatternFill("solid", fgColor="FFE699")
GREEN_FILL = PatternFill("solid", fgColor="C6E0B4")

wb = openpyxl.load_workbook(FPATH)

# --- helper column on Inventory_Planning_Model: running rank among Excess_Risk_SKU="Yes" rows ---
ipm = wb["Inventory_Planning_Model"]
ipm.cell(row=8, column=24, value="Excess_Risk_Rank")
ipm.cell(row=8, column=24).fill = HEADER_FILL
ipm.cell(row=8, column=24).font = HEADER_FONT
ipm.cell(row=8, column=24).alignment = Alignment(horizontal="center", wrap_text=True)
ipm.cell(row=8, column=24).border = BORDER
for r in range(9, 49):
    c = ipm.cell(row=r, column=24, value=f"=IF(V{r}=\"Yes\",COUNTIF($V$9:V{r},\"Yes\"),\"\")")
    c.font = FORMULA_FONT
    c.border = BORDER
ipm.column_dimensions["X"].width = 12

if "Executive_Summary" in wb.sheetnames:
    del wb["Executive_Summary"]
ws = wb.create_sheet("Executive_Summary", 2)  # place near the front
ws.sheet_view.showGridLines = False

ws["B2"] = "Meridian Manufacturing Co. — Inventory Health Executive Summary"
ws["B2"].font = TITLE_FONT
ws["B3"] = "40-SKU catalog | 12 months of sales & inventory history | All figures below are live formulas"
ws["B3"].font = SUBTITLE_FONT

# ---------------------------------------------------------------------------
# KPI CARDS (row 5-9)
# ---------------------------------------------------------------------------
kpis = [
    ("Total SKUs Analyzed", "=COUNTA(Inventory_Planning_Model!A9:A48)", "0"),
    ("SKUs at Stockout Risk", "=COUNTIF(Inventory_Planning_Model!R9:R48,\"Stockout Risk\")", "0"),
    ("SKUs Overstocked", "=COUNTIF(Inventory_Planning_Model!R9:R48,\"Overstock\")", "0"),
    ("Excess-Risk SKUs (Overstocked + Erratic)", "=COUNTIF(Inventory_Planning_Model!V9:V48,\"Yes\")", "0"),
    ("Total Avg. Inventory Value", "=SUM(Inventory_Planning_Model!W9:W48)", "$#,##0"),
    ("Value Tied Up in Excess-Risk SKUs", "=SUMIF(Inventory_Planning_Model!V9:V48,\"Yes\",Inventory_Planning_Model!W9:W48)", "$#,##0"),
    ("Avg. Inventory Turnover", "=AVERAGE(Inventory_Planning_Model!N9:N48)/AVERAGE(Inventory_Planning_Model!F9:F48*365)*365", "0.0\"x/yr\""),
    ("Forecast Accuracy (WAPE)", "=Forecast_MAPE_Summary!F6", "0.0%"),
]
# simpler, correct turnover: Annual COGS / Avg Inventory Value, both summed across catalog
kpis[6] = ("Catalog Inventory Turnover", "=SUMPRODUCT(Inventory_Planning_Model!K9:K48,Inventory_Planning_Model!E9:E48)/SUM(Inventory_Planning_Model!W9:W48)", "0.0\"x/yr\"")

col_positions = [2, 5, 8, 11]  # B, E, H, K -- 4 cards per row, 2 rows
row_starts = [5, 5, 5, 5, 9, 9, 9, 9]
row_pos = [5, 5, 5, 5, 9, 9, 9, 9]
for idx, (label, formula, fmt) in enumerate(kpis):
    row = 5 if idx < 4 else 9
    col = col_positions[idx % 4]
    for rr in range(row, row + 3):
        for cc in range(col, col + 2):
            ws.cell(row=rr, column=cc).fill = CARD_FILL
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = KPI_LABEL_FONT
    lbl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
    val = ws.cell(row=row + 1, column=col, value=formula)
    val.font = KPI_VALUE_FONT
    val.number_format = fmt
    val.alignment = Alignment(horizontal="left", vertical="center")

# ---------------------------------------------------------------------------
# ABC x XYZ MATRIX (row 13+)
# ---------------------------------------------------------------------------
ws.cell(row=13, column=2, value="ABC x XYZ Matrix — SKU Count").font = Font(name=FONT_NAME, bold=True, size=12, color="1F3864")
matrix_top = 15
cols_xyz = ["X (Steady)", "Y (Moderate)", "Z (Erratic)"]
rows_abc = ["A (Top 80% revenue)", "B (Next 15%)", "C (Bottom 5%)"]
ws.cell(row=matrix_top, column=2, value="")
for j, cx in enumerate(cols_xyz):
    c = ws.cell(row=matrix_top, column=3 + j, value=cx)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
for i, ra in enumerate(rows_abc):
    r = matrix_top + 1 + i
    lbl = ws.cell(row=r, column=2, value=ra)
    lbl.font = Font(name=FONT_NAME, bold=True, size=10)
    lbl.border = BORDER
    abc_letter = ra[0]
    for j, cx in enumerate(cols_xyz):
        xyz_letter = cx[0]
        cell = ws.cell(row=r, column=3 + j,
                        value=f"=COUNTIFS(ABC_XYZ_Classification!$H$6:$H$45,\"{abc_letter}\",ABC_XYZ_Classification!$L$6:$L$45,\"{xyz_letter}\")")
        cell.font = FORMULA_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
        if abc_letter == "A" and xyz_letter == "Z":
            cell.fill = RED_FILL
        elif xyz_letter == "Z":
            cell.fill = YELLOW_FILL
        elif xyz_letter == "X":
            cell.fill = GREEN_FILL

ws.cell(row=matrix_top + 5, column=2,
        value="AZ (top-revenue + erratic-demand) is the highest-attention cell: valuable, unpredictable, and expensive to get wrong either way.").font = Font(
    name=FONT_NAME, italic=True, size=9, color="7F7F7F")

# ---------------------------------------------------------------------------
# EXCESS-RISK SKU CALLOUT LIST (row ~22+)
# ---------------------------------------------------------------------------
callout_top = matrix_top + 8
ws.cell(row=callout_top, column=2, value="Excess-Risk SKUs (Overstocked + Erratic Demand)").font = Font(name=FONT_NAME, bold=True, size=12, color="1F3864")
headers = ["SKU_ID", "Description", "Category", "Days_of_Supply", "Lead_Time_Days", "Avg_Inventory_Value"]
hr = callout_top + 2
for j, h in enumerate(headers, start=2):
    c = ws.cell(row=hr, column=j, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER

# up to 8 rows -- pull the Nth "Yes" row from Inventory_Planning_Model using SMALL/IF-free approach:
# since LibreOffice here can't use array CSE reliably and array spill functions are unsupported,
# we instead pre-rank in a small helper column on Inventory_Planning_Model (added above) and pull via IFERROR+INDEX/MATCH per rank.
for k in range(8):
    r = hr + 1 + k
    rank = k + 1
    match_expr = f"MATCH({rank},Inventory_Planning_Model!$X$9:$X$48,0)"
    ws.cell(row=r, column=2,
            value=f"=IFERROR(INDEX(Inventory_Planning_Model!$A$9:$A$48,{match_expr}),\"—\")")
    ws.cell(row=r, column=3,
            value=f"=IFERROR(INDEX(Inventory_Planning_Model!$B$9:$B$48,{match_expr}),\"\")")
    ws.cell(row=r, column=4,
            value=f"=IFERROR(INDEX(Inventory_Planning_Model!$C$9:$C$48,{match_expr}),\"\")")
    ws.cell(row=r, column=5,
            value=f"=IFERROR(ROUND(INDEX(Inventory_Planning_Model!$P$9:$P$48,{match_expr}),0),\"\")")
    ws.cell(row=r, column=6,
            value=f"=IFERROR(INDEX(Inventory_Planning_Model!$D$9:$D$48,{match_expr}),\"\")")
    ws.cell(row=r, column=7,
            value=f"=IFERROR(ROUND(INDEX(Inventory_Planning_Model!$W$9:$W$48,{match_expr}),0),\"\")")
    for j in range(2, 8):
        ws.cell(row=r, column=j).font = FORMULA_FONT
        ws.cell(row=r, column=j).border = BORDER
        if j == 7:
            ws.cell(row=r, column=j).number_format = "$#,##0"

footer_row = hr + 11
ws.cell(row=footer_row, column=2,
        value="Source: simulated NetSuite-style export, Meridian Manufacturing Co. | Model built in Excel with live formulas | See README tab for methodology & assumptions."
        ).font = Font(name=FONT_NAME, italic=True, size=9, color="A6A6A6")

widths = {"A": 3, "B": 30, "C": 30, "D": 22, "E": 15, "F": 15, "G": 16, "H": 15, "I": 15, "J": 15, "K": 30, "L": 15}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(FPATH)
print("Stage 5 (Executive_Summary) saved.")
